"""
dump_xlsx.py — 把 xlsx 转成 UTF-8 的 csv / md / xlsx 三种格式,中文不乱码

xlsx 文件本质是 zip 里的 xml,openpyxl 读出的是 utf-8 字符串。
之前的 xlsx_view.py 错误地对已正确的字符串做 utf-8→gbk 反向 mojibake,
导致 dump 出来全是乱码。本脚本不做任何编码转换,直接以 utf-8 落盘。

用法:
  python tools/dump_xlsx.py data/石蜡结果+冰冻大模型.xlsx --out data/_xlsx_dump_v2
"""
from __future__ import annotations

import argparse
import csv
import json
import sys
from pathlib import Path

import openpyxl


def cell_value(v):
    """把 openpyxl 读到的值规整成可直接写入 csv/md 的字符串"""
    if v is None:
        return ""
    return v


def dump_sheet(ws, out_dir: Path, tag: str):
    """转储一个 sheet: csv + md + 追加到 xlsx(整本合一)"""
    rows = []
    for r in range(1, ws.max_row + 1):
        row = [cell_value(ws.cell(row=r, column=c).value) for c in range(1, ws.max_column + 1)]
        rows.append(row)

    # CSV
    with open(out_dir / f"{tag}.csv", "w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        w.writerows(rows)

    # MD (前若干行预览,避免大表)
    md = [f"# {tag}", "", f"原 sheet 名: **{ws.title}**  ", f"行数: {len(rows)}  列数: {len(rows[0]) if rows else 0}", ""]
    if rows:
        hdr = [str(c) for c in rows[0]]
        md.append("| " + " | ".join(hdr) + " |")
        md.append("|" + "|".join("---" for _ in hdr) + "|")
        for r in rows[1:51]:  # 只截前 50 行
            md.append("| " + " | ".join(
                str(c).replace("\n", " ").replace("|", r"\|")[:160] for c in r
            ) + " |")
        if len(rows) > 51:
            md.append(f"\n_(省略剩余 {len(rows) - 51} 行,见 csv)_")
    (out_dir / f"{tag}.md").write_text("\n".join(md), encoding="utf-8")

    # JSON(整表)
    (out_dir / f"{tag}.json").write_text(
        json.dumps(rows, ensure_ascii=False, indent=2, default=str),
        encoding="utf-8",
    )

    return rows


def main() -> int:
    ap = argparse.ArgumentParser(description=__doc__.splitlines()[0])
    ap.add_argument("xlsx", help="input .xlsx file")
    ap.add_argument("--out", required=True, help="output directory")
    args = ap.parse_args()

    src = Path(args.xlsx)
    out = Path(args.out)
    out.mkdir(parents=True, exist_ok=True)

    wb_src = openpyxl.load_workbook(src, data_only=True)
    # 写一个 utf-8 的 xlsx(供 Excel/WPS 直接打开)
    wb_out = openpyxl.Workbook()
    wb_out.remove(wb_out.active)

    print(f"[src] {src} ({src.stat().st_size} bytes)")
    print(f"[src] sheets: {wb_src.sheetnames}")

    summary = []
    for idx, ws in enumerate(wb_src.worksheets):
        tag = f"{idx+1:02d}_{ws.title}"
        # tag 安全化(去路径危险字符)
        safe_tag = "".join(ch if ch.isalnum() or ch in "_一-鿿" else "_" for ch in tag)[:60]
        n = len(dump_sheet(ws, out, safe_tag))
        # 复制到 out xlsx
        ws_out = wb_out.create_sheet(title=ws.title[:31])  # excel 限 31 字符
        for r in range(1, ws.max_row + 1):
            for c in range(1, ws.max_column + 1):
                ws_out.cell(row=r, column=c, value=ws.cell(row=r, column=c).value)
        summary.append((safe_tag, n, ws.max_column))
        print(f"  [{idx}] '{ws.title}' -> {safe_tag} (rows={n}, cols={ws.max_column})")

    out_xlsx = out / f"{src.stem}_utf8.xlsx"
    wb_out.save(out_xlsx)
    print(f"\n[done] wrote {len(summary)} sheets + {out_xlsx.name} to {out}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
