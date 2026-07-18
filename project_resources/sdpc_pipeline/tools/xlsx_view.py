"""
xlsx_view.py — [已废弃] 早期 xlsx dump 工具,会把正常中文做反向 mojibake

**请改用 tools/dump_xlsx.py** —— 它不进行任何编码转换,直接以 UTF-8 落盘。

本文件保留仅为历史参考;不要再用 `python tools/xlsx_view.py` 生成 dump。

from __future__ import annotations

import argparse
import csv
import json
import sys
from pathlib import Path

import openpyxl


def try_fix_mojibake(s: str) -> tuple[str, bool]:
    """Try utf-8 -> gbk round trip. Return (decoded, success)."""
    if not isinstance(s, str):
        return s, False
    try:
        return s.encode("utf-8").decode("gbk"), True
    except Exception:
        return s, False


def fix_workbook(xlsx_path: str) -> list[dict]:
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    out = []
    for idx, ws in enumerate(wb.worksheets):
        sheet_info = {
            "index": idx,
            "raw_name": ws.title,
            "decoded_name": try_fix_mojibake(ws.title)[0],
            "rows": ws.max_row,
            "cols": ws.max_column,
            "data": [],
        }
        for r in range(1, ws.max_row + 1):
            row = []
            for c in range(1, ws.max_column + 1):
                v = ws.cell(row=r, column=c).value
                if isinstance(v, str):
                    decoded, ok = try_fix_mojibake(v)
                    row.append({"raw": v, "decoded": decoded, "fixed": ok})
                else:
                    row.append({"raw": v, "decoded": v, "fixed": False})
            sheet_info["data"].append(row)
        out.append(sheet_info)
    return out


def main() -> int:
    ap = argparse.ArgumentParser(description=__doc__.splitlines()[0])
    ap.add_argument("xlsx", help="path to .xlsx file")
    ap.add_argument("--sheet", type=int, help="0-based sheet index to print")
    ap.add_argument("--out", type=str, help="output dir for fixed JSON+CSV+MD")
    args = ap.parse_args()

    sheets = fix_workbook(args.xlsx)
    for s in sheets:
        print(f"[{s['index']}] raw={s['raw_name']!r}  decoded={s['decoded_name']!r}  "
              f"rows={s['rows']} cols={s['cols']}")

    if args.sheet is not None:
        s = sheets[args.sheet]
        print(f"\n--- Sheet {s['index']} (decoded: {s['decoded_name']}) ---")
        for r in s["data"][:30]:
            cells = [c["decoded"] if c["decoded"] is not None else "" for c in r]
            print("  " + " | ".join(str(x)[:60] for x in cells))

    if args.out:
        out = Path(args.out)
        out.mkdir(exist_ok=True, parents=True)
        for s in sheets:
            tag = s["decoded_name"] or f"sheet{s['index']}"
            tag = "".join(ch if ch.isalnum() or ch in "_+" else "_" for ch in tag)[:60]
            tag = f"{s['index']+1:02d}_{tag}"
            # JSON: keep raw + decoded
            (out / f"{tag}.json").write_text(
                json.dumps(s, ensure_ascii=False, indent=2, default=str),
                encoding="utf-8",
            )
            # CSV: decoded only
            with open(out / f"{tag}.csv", "w", newline="", encoding="utf-8") as f:
                w = csv.writer(f)
                for r in s["data"]:
                    w.writerow([c["decoded"] if c["decoded"] is not None else "" for c in r])
            # MD: decoded only
            if s["data"]:
                hdr = [c["decoded"] if c["decoded"] is not None else "" for c in s["data"][0]]
                md = [f"# {tag}",
                      "",
                      f"原 sheet 名 (raw): `{s['raw_name']}`  ",
                      f"解码后: **{s['decoded_name']}**  ",
                      f"行数: {s['rows']}  列数: {s['cols']}",
                      "",
                      "| " + " | ".join(str(h) for h in hdr) + " |",
                      "|" + "|".join("---" for _ in hdr) + "|"]
                for r in s["data"][1:]:
                    md.append("| " + " | ".join(
                        (str(c["decoded"])[:120].replace("\n", " ") if c["decoded"] is not None else "")
                        for c in r) + " |")
                (out / f"{tag}.md").write_text("\n".join(md), encoding="utf-8")
        print(f"\nWrote {len(sheets)} sheets to {out}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
